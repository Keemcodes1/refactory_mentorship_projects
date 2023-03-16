import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def get_product_data(filename):
    """Reads product data from an Excel file and returns a dictionary."""
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    products = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        product, price = row
        products[product] = price

    return products


def create_account():
    """Prompts the buyer to create an account and saves it to an Excel file."""
    email = input("Enter your email: ")
    password = input("Enter your password: ")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value="Email")
    worksheet.cell(row=1, column=2, value="Password")

    worksheet.cell(row=2, column=1, value=email)
    worksheet.cell(row=2, column=2, value=password)

    workbook.save("accounts.xlsx")


def login():
    """Prompts the buyer to login and returns True if successful."""
    email = input("Enter your email: ")
    password = input("Enter your password: ")

    workbook = openpyxl.load_workbook("accounts.xlsx")
    worksheet = workbook.active

    if worksheet.cell(row=2, column=1).value == email and worksheet.cell(row=2, column=2).value == password:
        return True
    else:
        return False


def buy(products):
    """Prompts the buyer to select products to purchase and returns the total price."""
    total = 0

    while True:
        print("Available Products:")
        for product, price in products.items():
            print(f"{product} - ${price}")

        selection = input("Enter the name of the product you want to buy or 'done' to finish: ")

        if selection == "done":
            break

        if selection not in products:
            print("Invalid selection.")
            continue

        quantity = input(f"How many {selection} do you want to buy? ")

        try:
            quantity = int(quantity)
        except ValueError:
            print("Invalid quantity.")
            continue

        total += products[selection] * quantity

    return total


def process_payment(total):
    """Prompts the buyer to pay and returns True if the payment is successful."""
    while True:
        amount = input(f"Enter the amount you want to pay (${total} due): ")

        try:
            amount = float(amount)
        except ValueError:
            print("Invalid amount.")
            continue

        if amount < total:
            print("Payment amount is less than the total amount due.")
            continue

        change = amount - total
        print(f"Change: ${change:.2f}")
        return True


def print_receipt(email, total):
    """Prints a receipt to the console."""
    print("\nReceipt:")
    print("=" * 30)
    print(f"Buyer: {email}")
    print(f"Total: ${total:.2f}")
    print("=" * 30)


def main():
    # Get product data from Excel file
    products = get_product_data("products.xlsx")

    # Prompt buyer to create account or login
    if login():
        email = input("Enter your email: ")
    else:
        create_account()
        email = input("Enter your email: ")

    # Buy products and process payment
    total = buy(products)
    if process_payment(total):
        print_receipt(email, total)


if __name__ == "__main__":
    main()

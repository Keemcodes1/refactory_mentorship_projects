
import openpyxl

# Open the Excel file
workbook = openpyxl.load_workbook('shop.xlsx')

# Get the accounts sheet
accounts_sheet = workbook['accounts']

# Prompt user to enter email and password
email = input("Enter your email: ")
password = input("Enter your password: ")

# Add user's email and password to the accounts sheet
accounts_sheet.append([email, password])

# Save the Excel file
workbook.save('shop.xlsx')



def login():
    # Open the Excel file
    workbook = openpyxl.load_workbook('shop.xlsx')

    # Get the accounts sheet
    accounts_sheet = workbook['accounts']

    # Prompt user to enter email and password
    email = input("Enter your email: ")
    password = input("Enter your password: ")

    # Check if email and password exist in the accounts sheet
    for row in accounts_sheet.iter_rows(values_only=True):
        if row[0] == email and row[1] == password:
            print("You have successfully logged in!")
            return True

    # If email and password do not match any records, prompt user to try again
    print("Incorrect email or password. Please try again.")
    return False



def view_products():
    # Open the Excel file
    workbook = openpyxl.load_workbook('shop.xlsx')

    # Get the products sheet
    products_sheet = workbook['products']

    # Print the list of available products and their prices
    print("Available products:")
    for row in products_sheet.iter_rows(values_only=True):
        print(f"{row[0]} - {row[1]}")



def buy_products():
    # Open the Excel file
    workbook = openpyxl.load_workbook('shop.xlsx')

    # Get the products sheet
    products_sheet = workbook['products']

    # Prompt user to enter product name and quantity
    product_name = input("Enter the name of the product you want to buy: ")
    quantity = int(input("Enter the quantity you want to buy: "))

    # Find the price of the product
    for row in products_sheet.iter_rows(values_only=True):
        if row[0] == product_name:
            price = row[1]
            break

    # Calculate the total price of the purchase
    total_price = price * quantity

    # Prompt user to pay for the purchase
    while True:
        payment = int(input(f"The total price of your purchase is {total_price}. Enter amount paid: "))
        if payment < total_price:
            print("Payment is less than the total price. Please try again.")
        else:
            break

    # Print the receipt
    print("----- Receipt -----")
    print(f"Product: {product_name}")
    print(f"Quantity: {quantity}")
    print(f"Price per unit: {price}")
    print(f"Total price: {total_price}")
    print(f"Amount paid: {payment}")
    print(f"Change: {payment - total_price}")



def main():
    while True:
        print("1. Create account")
        print("2. Login")
        print("3. View products")
        print("4. Buy products")
        print("5. Exit")
        choice = int(input("Enter your choice: "))

        if choice == 1:
            create_account()
        elif choice == 2:
            if login():
                print("You can now view or buy products.")
            else:
                print("Please create an account or try again.")
        elif choice == 3:
            view_products()
        elif choice == 4:
            buy_products()
        elif choice == 5:
            print("Thank you for shopping with us!")
            break
        else:
            print("Invalid choice. Please try again.")




import openpyxl
import xlwt

# Set up the workbook and sheet
keem = xlwt.Workbook()
sheet1 = keem.add_sheet("my excel sheet")

# Write the product information to the sheet
sheet1.write(0, 0, 'Product')
sheet1.write(0, 1, 'Price')
sheet1.write(1, 0, 'laptops')
sheet1.write(1, 1, 0.99)
sheet1.write(2, 0, 'phones')
sheet1.write(2, 1, 0.25)
sheet1.write(3, 0, 'watches')
sheet1.write(3, 1, 0.50)

# Write the account information to the sheet
sheet1.write(5, 0, 'Email')
sheet1.write(5, 1, 'Password')
sheet1.write(6, 0, 'test@example.com')
sheet1.write(6, 1, 'password123')

# Saving the workbook to a file
keem.save("products.xls")

# Load the product information from the sheet and convert to a dictionary
products = {}
for row in range(1, 4):
    
    workbook = openpyxl.load_workbook("products.xls")
    worksheet = workbook.get_sheet_by_name("my excel sheet")
    name = sheet1['A']
    price = sheet1.cell(row, 1).value
    products[name] = price

# Load the account information from the sheet and convert to a dictionary
accounts = {}
for row in range(6, 7):
    email = sheet1.cell(row, 0).value
    password = sheet1.cell(row, 1).value
    accounts[email] = password

# Prompt the buyer to create an account
while True:
    email = input("Enter your email address: ")
    if email in accounts:
        print("An account with that email address already exists. Please try again.")
        continue
    else:
        password = input("Choose a password: ")
        accounts[email] = password
        sheet1.write(len(accounts) + 5, 0, email)
        sheet1.write(len(accounts) + 5, 1, password)
        keem.save("products.xls")
        break

# Prompt the buyer to login
while True:
    email_input = input("Enter your email address: ")
    password_input = input("Enter your password: ")

    if email_input in accounts and accounts[email_input] == password_input:
        break
    else:
        print("Invalid email or password. Please try again.")

# Show the available products
print("Available products:")
for name, price in products.items():
    print(f"{name}: ${price}")

# Let the buyer choose what to buy
cart = {}
while True:
    item = input("Enter a product name to add to your cart, or 'done' to checkout: ")
    if item == 'done':
        break

    if item not in products:
        print("Invalid product name. Please try again.")
        continue

    quantity = int(input(f"How many {item}s would you like to buy? "))
    cart[item] = quantity

# Calculate the total price
total_price = sum(products[item] * quantity for item, quantity in cart.items())

# Ask the buyer to pay
while True:
    payment = float(input(f"Your total is ${total_price}. Enter your payment amount: "))
    if payment < total

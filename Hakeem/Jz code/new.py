from openpyxl import Workbook, load_workbook

# Define the paths to the Excel files
ACCOUNTS_PATH = 'accounts.xlsx'
PRODUCTS_PATH = 'products.xlsx'

# Load the accounts from the Excel sheet
accounts = {}
try:
    accounts_wb = load_workbook(ACCOUNTS_PATH)
    accounts_ws = accounts_wb.active
    for row in accounts_ws.iter_rows(values_only=True):
        email, password = row
        accounts[email] = password
except FileNotFoundError:
    accounts_wb = Workbook()
    accounts_ws = accounts_wb.active

# Load the products from the Excel sheet
products_wb = load_workbook(PRODUCTS_PATH)
products_ws = products_wb.active
products = []
for row in products_ws.iter_rows(values_only=True):
    product, price = row
    products.append({'Product': product, 'Price': price})

# Define a function to save the accounts to the Excel sheet
def save_accounts():
    accounts_ws.delete_rows(1, accounts_ws.max_row)
    for i, (email, password) in enumerate(accounts.items(), start=1):
        accounts_ws.cell(row=i, column=1, value=email)
        accounts_ws.cell(row=i, column=2, value=password)
    accounts_wb.save(ACCOUNTS_PATH)

# Define a function to create a new account for the buyer
def create_account():
    email = input("Enter your email address: ")
    password = input("Create a password: ")
    accounts[email] = password
    save_accounts()
    print("Account created successfully!")

# Define a function to ask the buyer to log in
def log_in():
    email = input("Enter your email address: ")
    password = input("Enter your password: ")
    if email in accounts and accounts[email] == password:
        print("Log in successful!")
    else:
        print("Incorrect email or password. Please try again.")
        log_in()

# Define a function to display the list of available products
def display_products():
    print("Available products:")
    for product in products:
        print(f"{product['Product']}: ${product['Price']}")

# Define a function to let the buyer choose what to buy
def buy_products():
    total_price = 0.0
    cart = []
    while True:
        product = input("Enter the name of the product you want to buy (or 'done' to finish): ")
        if product == 'done':
            break
        product_found = False
        for p in products:
            if p['Product'] == product:
                product_found = True
                price = float(p['Price'])
                total_price += price
                cart.append((product, price))
                print(f"{product} added to cart. Current total price: ${total_price:.2f}")
                break
        if not product_found:
            print("Invalid product. Please try again.")
    return cart, total_price

# Define a function to print a receipt for the buyer
def print_receipt(cart, total_price):
    print("------- Receipt -------")
    for product, price in cart:
        print(f"{product}: ${price:.2f}")
    print(f"Total: ${total_price:.2f}")

# Prompt the buyer to create an account or log in
if input("Do you have an account? (y/n) ").lower() == 'n':
    create_account()
log_in()

# Display